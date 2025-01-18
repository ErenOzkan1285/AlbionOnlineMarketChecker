import openpyxl
from openpyxl.styles import Alignment, PatternFill
import os
import re

# File names
output_file = "market_prices.xlsx"
txt_file = "market_prices.txt"

# Cities and their column structure
cities = ["Caerleon", "Lymhurst", "Martlock", "Thetford", "Fort Sterling", "Bridgewatch"]
columns = ["Name", "Tier", "Enchantment", "Quality", "Price"]

# Define colors for each city
city_colors = {
    "Caerleon": None,           # No color
    "Lymhurst": "00FF00",       # Green
    "Martlock": "0000FF",       # Blue
    "Thetford": "800080",       # Purple
    "Fort Sterling": "C0C0C0",  # Grey
    "Bridgewatch": "FFA500",    # Orange
}

# Cities and their corresponding codes
city_codes = {
    "C": "Caerleon",
    "L": "Lymhurst",
    "M": "Martlock",
    "T": "Thetford",
    "F": "Fort Sterling",
    "B": "Bridgewatch"
}

def create_xlsx_schema():
    # Create a new workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Market Data"

    # Start writing headers
    start_col = 1
    for city in cities:
        # Write city name merged over the columns
        end_col = start_col + len(columns) - 1
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = sheet.cell(row=1, column=start_col)
        cell.value = city
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply background color to the city header
        city_color = city_colors.get(city)
        if city_color:
            fill = PatternFill(start_color=city_color, end_color=city_color, fill_type="solid")
            cell.fill = fill

        # Write column headers
        for i, col in enumerate(columns):
            sheet.cell(row=2, column=start_col + i, value=col)

        # Move to the next city block
        start_col = end_col + 2

    # Save the workbook
    workbook.save(output_file)
    print(f"File '{output_file}' created successfully.")

# Function to parse the text file and extract details
def parse_txt_file():
    with open(txt_file, "r") as file:
        lines = file.readlines()

    # Variable to store parsed data
    item_data = []

    i = 0
    while i < len(lines):
        # Search for the item details line (e.g., "Item Details - Tier: 6, Enchantment: 1, Quality: 0, City: L")
        item_details = re.match(r"Item Details - Tier: (\d), Enchantment: (\d), Quality: (\d), City: (\w)", lines[i].strip())
        if item_details:
            # Extract tier, enchantment, quality, and city
            tier, enchantment, quality, city_code = item_details.groups()
            city_name = city_codes.get(city_code, "Unknown City")

            # Move to the next line for item names and prices
            i += 1
            items = []
            prices = []

            while i < len(lines) and not lines[i].startswith("Item Details"):
                # Collect item names
                if lines[i].strip() and not re.match(r"^\d{1,3}(?:,\d{3})*$", lines[i].strip()):
                    items.append(lines[i].strip())
                elif re.match(r"^\d{1,3}(?:,\d{3})*$", lines[i].strip()):
                    # Collect corresponding prices
                    price = int(lines[i].strip().replace(",", ""))
                    prices.append(price)
                i += 1

            # Ensure that items and prices are in pairs, and append them
            for item, price in zip(items, prices):
                item_data.append((tier, enchantment, quality, city_name, item, price))
        else:
            i += 1

    return item_data

# Function to append parsed data to the Excel sheet
def append_to_excel(item_data):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    # Start column index
    start_col = 1
    for city in city_codes.values():
        end_col = start_col + 4  # 5 columns per city (Name, Tier, Enchantment, Quality, Price)

        # Find the first empty row for the city
        first_empty_row = sheet.max_row + 1
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=start_col).value is None:
                first_empty_row = row
                break

        # Append the data for each city
        for row in item_data:
            if row[3] == city:  # Check if the city matches
                sheet.cell(row=first_empty_row, column=start_col, value=row[4])  # Item Name
                sheet.cell(row=first_empty_row, column=start_col + 1, value=row[0])  # Tier
                sheet.cell(row=first_empty_row, column=start_col + 2, value=row[1])  # Enchantment
                sheet.cell(row=first_empty_row, column=start_col + 3, value=row[2])  # Quality
                sheet.cell(row=first_empty_row, column=start_col + 4, value=row[5])  # Price
                first_empty_row += 1  # Move to the next row for the next item

        start_col = end_col + 2  # Move to the next city block

    # Save the workbook after appending data
    workbook.save(output_file)
    print(f"Data successfully appended to '{output_file}'.")

# Main function to process the data
def main():
    # Check if the file exists
    if not os.path.exists(output_file):
        create_xlsx_schema()  # Create the schema if the file doesn't exist
    
    item_data = parse_txt_file()  # Parse the txt file
    append_to_excel(item_data)    # Append parsed data to Excel sheet

# Run the program
if __name__ == "__main__":
    main()  # Process and append the data
